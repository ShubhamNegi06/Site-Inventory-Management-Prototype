"""
Parses an uploaded .xlsx against the sample template's two-row header
(section row + field-label row -- see seed_fields.py / lib/sections.ts on
the frontend for the same convention) and turns it into rows shaped like
what SampleCreate expects, WITHOUT touching the database. Validation only
-- the actual insert happens in the /bulk-import/commit route, one row at a
time, after the frontend has shown the user a preview and they've
confirmed.

Column headers are matched against FieldDefinition.field_label (not
field_key), because field_label is the human-readable text that actually
appears in the spreadsheet -- the same lookup the "add sample" form's
labels come from. This means the mapping self-updates as new fields get
added via the UI, with no separate maintenance.
"""
import io
from datetime import date, datetime
from typing import Any

import openpyxl
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.models.field_definition import FieldDefinition
from app.models.sample import Sample
from app.models.user import User, UserRole
from app.schemas.sample import BulkImportRow, BulkImportPreviewResponse

# Header text (normalized) that maps onto Sample's own fixed columns,
# rather than becoming a key in the JSONB `data` blob.
FIXED_COLUMN_MAP = {
    "subject id": "subject_code",
    "sample id": "sample_code",
    "sample type": "sample_type",
    "date of sample collection": "collection_date",
}

# Present in the template purely as a row counter -- never stored.
IGNORED_HEADERS = {"sr. no.", "sr no.", "sr no", "s.no", "s. no.", "sno"}


def _normalize_header(value: Any) -> str:
    """Collapses whitespace/newlines and lowercases, so "Type of Tissue "
    and "Type of Tissue" (or a header wrapped across two lines in the
    sheet) compare equal."""
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip().lower()


def _normalize_cell(value: Any) -> Any:
    """Cleans a single data-column cell for storage: trims strings, and
    turns openpyxl's native date/datetime objects into ISO strings so they
    round-trip through JSON the same way the "add sample" form's date
    inputs do. Anything else (numbers, bools) passes through unchanged."""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _parse_date_cell(value: Any) -> tuple[date | None, bool]:
    """Returns (parsed_date, ok). ok=False means the cell had *something*
    in it that couldn't be read as a date, so the caller can flag it --
    distinct from an empty cell, which is just None/True."""
    if value in (None, ""):
        return None, True
    if isinstance(value, datetime):
        return value.date(), True
    if isinstance(value, date):
        return value, True
    text = str(value).strip()
    if not text:
        return None, True
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date(), True
        except ValueError:
            continue
    return None, False


def _find_header_row(ws) -> int | None:
    """
    The template has a merged "section" row (Case Details, Demographic
    Details, ...) sitting above the actual column headers, so the row we
    want isn't a fixed index -- it's whichever row contains both "Subject
    ID" and "Sample ID". Scanning the first few rows for that combination
    keeps this working even if a section row is missing, extra, or
    reordered.
    """
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = {_normalize_header(c.value) for c in row}
        if "subject id" in values and "sample id" in values:
            return row[0].row
    return None


def _build_label_lookup(db: Session, user: User) -> dict[str, str]:
    """normalized field_label -> field_key, scoped exactly like the "add
    sample" form's own field list: every global field, plus (for a site
    user) that site's own custom fields."""
    query = db.query(FieldDefinition)
    if user.role == UserRole.site:
        query = query.filter(
            or_(FieldDefinition.site_id.is_(None), FieldDefinition.site_id == user.site_id)
        )
    return {_normalize_header(f.field_label): f.field_key for f in query.all()}


def parse_workbook(content: bytes, db: Session, user: User) -> BulkImportPreviewResponse:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise ValueError(f"Couldn't read this as an Excel file ({e}). Please upload a .xlsx.")

    label_lookup = _build_label_lookup(db, user)
    existing_codes = {
        code for (code,) in db.query(Sample.sample_code).filter(Sample.is_deleted.is_(False)).all()
    }
    seen_in_file: set[str] = set()
    unmapped: set[str] = set()
    rows: list[BulkImportRow] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row_idx = _find_header_row(ws)
        if header_row_idx is None:
            continue  # this tab doesn't look like a sample table -- skip it, not an error

        headers = [c.value for c in ws[header_row_idx]]
        # (column_index, "fixed" | "data", target_key)
        col_map: list[tuple[int, str, str]] = []
        for idx, raw_header in enumerate(headers):
            norm = _normalize_header(raw_header)
            if not norm or norm in IGNORED_HEADERS:
                continue
            if norm in FIXED_COLUMN_MAP:
                col_map.append((idx, "fixed", FIXED_COLUMN_MAP[norm]))
            elif norm in label_lookup:
                col_map.append((idx, "data", label_lookup[norm]))
            else:
                unmapped.add(str(raw_header).strip())

        for row_cells in ws.iter_rows(min_row=header_row_idx + 1):
            values = [c.value for c in row_cells]
            if all(v in (None, "") for v in values):
                continue  # blank row -- common at the end of a sheet

            row_number = row_cells[0].row
            fixed: dict[str, Any] = {}
            data: dict[str, Any] = {}
            for idx, kind, key in col_map:
                if idx >= len(values):
                    continue
                (fixed if kind == "fixed" else data)[key] = values[idx]

            errors: list[str] = []

            def clean_str(v: Any) -> str | None:
                if v in (None, ""):
                    return None
                text = str(v).strip()
                return text or None

            subject_code = clean_str(fixed.get("subject_code"))
            sample_code = clean_str(fixed.get("sample_code"))
            sample_type = clean_str(fixed.get("sample_type"))

            if not subject_code:
                errors.append("Missing Subject ID")
            if not sample_code:
                errors.append("Missing Sample ID")
            if not sample_type:
                errors.append("Missing Sample Type")

            if sample_code:
                if sample_code in existing_codes:
                    errors.append(f'Sample ID "{sample_code}" already exists in the inventory')
                elif sample_code in seen_in_file:
                    errors.append(f'Sample ID "{sample_code}" is duplicated elsewhere in this file')
                else:
                    seen_in_file.add(sample_code)

            collection_date, date_ok = _parse_date_cell(fixed.get("collection_date"))
            if not date_ok:
                errors.append(f'Could not read "Date of Sample Collection" value: {fixed.get("collection_date")!r}')

            cleaned_data = {key: _normalize_cell(value) for key, value in data.items() if value not in (None, "")}

            rows.append(
                BulkImportRow(
                    row_number=row_number,
                    sheet=sheet_name,
                    subject_code=subject_code,
                    sample_code=sample_code,
                    sample_type=sample_type,
                    collection_date=collection_date,
                    data=cleaned_data,
                    errors=errors,
                )
            )

    valid_count = sum(1 for r in rows if not r.errors)
    return BulkImportPreviewResponse(
        rows=rows,
        unmapped_columns=sorted(unmapped),
        valid_count=valid_count,
        error_count=len(rows) - valid_count,
    )
